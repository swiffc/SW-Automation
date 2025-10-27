"""
Excel Design Table Structure Analyzer
Reads Excel files and maps all parameters for UI generation
"""
# -*- coding: utf-8 -*-

import openpyxl
import json
from pathlib import Path

def analyze_excel_file(filepath):
    """Analyze an Excel file and extract structure"""
    
    print(f"\n{'='*60}")
    print(f"Analyzing: {Path(filepath).name}")
    print(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        result = {
            'filename': Path(filepath).name,
            'sheets': []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"\nSheet: {sheet_name}")
            print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
            
            sheet_data = {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'parameters': [],
                'formulas': []
            }
            
            # Scan first 50 rows to find structure
            print("\nScanning first 20 rows:")
            for row in range(1, min(21, ws.max_row + 1)):
                row_data = []
                has_data = False
                
                for col in range(1, min(15, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    value = cell.value
                    
                    if value:
                        has_data = True
                        
                        # Check if it's a formula
                        if isinstance(value, str) and value.startswith('='):
                            sheet_data['formulas'].append({
                                'cell': f'{cell.column_letter}{cell.row}',
                                'formula': value
                            })
                            row_data.append(f"[FORMULA]")
                        else:
                            row_data.append(str(value)[:30])  # Truncate long values
                    else:
                        row_data.append("")
                
                if has_data:
                    print(f"Row {row:2d}: {' | '.join(row_data)}")
            
            result['sheets'].append(sheet_data)
        
        return result
        
    except Exception as e:
        print(f"ERROR: {e}")
        return None

def main():
    """Analyze all master config files"""
    
    print("\n" + "="*60)
    print("EXCEL DESIGN TABLE STRUCTURE ANALYZER")
    print("="*60)
    
    # Define files to analyze
    files_to_analyze = [
        "templates/header_section_tool/Combined_/Drafting/Headers/000000_S01c-HCS.xlsx",
        "templates/header_section_tool/Single_/Drafting/Headers/000000_S03-HCS.xlsx",
        "templates/xch_structure_tool/XCH Cooler/XCH_SCS.xlsx",
    ]
    
    results = []
    
    for filepath in files_to_analyze:
        full_path = Path(filepath)
        if full_path.exists():
            result = analyze_excel_file(full_path)
            if result:
                results.append(result)
        else:
            print(f"\n??  File not found: {filepath}")
    
    # Save results to JSON
    output_file = "docs/excel_structure_analysis.json"
    with open(output_file, 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n\n{'='*60}")
    print(f"? Analysis complete!")
    print(f"Results saved to: {output_file}")
    print(f"{'='*60}\n")
    
    # Print summary
    print("\nSUMMARY:")
    for result in results:
        print(f"\n?? {result['filename']}")
        for sheet in result['sheets']:
            print(f"   Sheet: {sheet['name']}")
            print(f"   Size: {sheet['max_row']} rows x {sheet['max_column']} columns")
            print(f"   Formulas found: {len(sheet['formulas'])}")

if __name__ == "__main__":
    main()

