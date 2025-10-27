"""
Excel Template Reader
Reads S25140-Prego1.xlsm to understand structure, formulas, and data
for implementing auto-populate functionality in UnifiedUI
"""

import openpyxl
from openpyxl import load_workbook
import json

def read_excel_template(file_path):
    """Read all sheets, cells, formulas, and data from Excel file"""
    
    print(f"Reading Excel file: {file_path}")
    print("=" * 80)
    
    # Load workbook
    wb = load_workbook(file_path, data_only=False)  # Keep formulas
    
    template_data = {
        'filename': file_path.split('\\')[-1],
        'sheets': []
    }
    
    # Read each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'=' * 80}")
        
        ws = wb[sheet_name]
        
        sheet_data = {
            'name': sheet_name,
            'dimensions': f"{ws.dimensions}",
            'cells': [],
            'formulas': [],
            'named_ranges': []
        }
        
        # Read all cells with data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_info = {
                        'address': cell.coordinate,
                        'value': str(cell.value)[:100],  # Limit length
                        'data_type': str(cell.data_type)
                    }
                    
                    # Check if it's a formula
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_info['formula'] = cell.value
                        sheet_data['formulas'].append(cell_info)
                        print(f"  FORMULA [{cell.coordinate}]: {cell.value[:80]}")
                    else:
                        # Regular value
                        if not isinstance(cell.value, str) or len(cell.value.strip()) > 0:
                            sheet_data['cells'].append(cell_info)
                    
                    # Show important cells (labels and values)
                    if cell.column <= 5 and cell.row <= 100:  # First few columns/rows
                        print(f"  {cell.coordinate}: {str(cell.value)[:60]}")
        
        print(f"\n  Total cells with data: {len(sheet_data['cells'])}")
        print(f"  Total formulas: {len(sheet_data['formulas'])}")
        
        template_data['sheets'].append(sheet_data)
    
    # Save to JSON for analysis
    output_file = 'excel_template_analysis.json'
    with open(output_file, 'w') as f:
        json.dump(template_data, f, indent=2)
    
    print(f"\n{'=' * 80}")
    print(f"Analysis saved to: {output_file}")
    print(f"Total sheets: {len(template_data['sheets'])}")
    
    return template_data


def create_summary_report(template_data):
    """Create a readable summary report"""
    
    report_file = 'EXCEL_TEMPLATE_SUMMARY.md'
    
    with open(report_file, 'w') as f:
        f.write(f"# Excel Template Analysis: {template_data['filename']}\n\n")
        f.write("## Summary\n\n")
        f.write(f"**Total Sheets:** {len(template_data['sheets'])}\n\n")
        
        for sheet in template_data['sheets']:
            f.write(f"\n## Sheet: {sheet['name']}\n\n")
            f.write(f"**Dimensions:** {sheet['dimensions']}\n\n")
            f.write(f"**Total Data Cells:** {len(sheet['cells'])}\n\n")
            f.write(f"**Total Formulas:** {len(sheet['formulas'])}\n\n")
            
            if sheet['formulas']:
                f.write(f"\n### Key Formulas:\n\n")
                for formula in sheet['formulas'][:20]:  # First 20 formulas
                    f.write(f"- **{formula['address']}**: `{formula['formula'][:100]}`\n")
                
                if len(sheet['formulas']) > 20:
                    f.write(f"\n... and {len(sheet['formulas']) - 20} more formulas\n")
    
    print(f"Summary report saved to: {report_file}")


if __name__ == '__main__':
    file_path = r'c:\Users\DCornealius\CascadeProjects\Solidworks_Automation\S25140-Prego1.xlsm'
    
    try:
        template_data = read_excel_template(file_path)
        create_summary_report(template_data)
        print("\n✅ Excel template analysis complete!")
        
    except Exception as e:
        print(f"\n❌ Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()
