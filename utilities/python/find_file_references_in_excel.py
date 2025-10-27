# -*- coding: utf-8 -*-
"""
Find File References in Excel Design Tables
Scans Excel files for references to CAD files that may need updating
"""

import openpyxl
from pathlib import Path
import re

def find_cad_references(filepath):
    """Find all references to CAD files in Excel"""
    
    print(f"\n{'='*70}")
    print(f"Scanning: {Path(filepath).name}")
    print(f"{'='*70}")
    
    references_found = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            
            # Scan all cells for CAD file references
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    
                    if value and isinstance(value, str):
                        # Look for SolidWorks file extensions
                        if any(ext in value.upper() for ext in ['.SLDASM', '.SLDPRT', '.SLDDRW', '.SLDPRT', '.SLDASM']):
                            references_found.append({
                                'sheet': sheet_name,
                                'cell': f'{cell.column_letter}{cell.row}',
                                'value': value,
                                'type': 'direct_reference'
                            })
                            print(f"  Found: {cell.column_letter}{cell.row} = {value[:60]}")
                        
                        # Look for file names that might need prefixes
                        # JOBNO pattern (Hudson)
                        if 'JOBNO' in value.upper():
                            references_found.append({
                                'sheet': sheet_name,
                                'cell': f'{cell.column_letter}{cell.row}',
                                'value': value,
                                'type': 'jobno_reference',
                                'needs_prefix': 'HUD_'
                            })
                            print(f"  JOBNO: {cell.column_letter}{cell.row} = {value[:60]}")
                        
                        # Z_ pattern (Z Tool)
                        if value.startswith('Z_'):
                            references_found.append({
                                'sheet': sheet_name,
                                'cell': f'{cell.column_letter}{cell.row}',
                                'value': value,
                                'type': 'z_reference',
                                'needs_prefix': 'ZST_'
                            })
                            print(f"  Z_: {cell.column_letter}{cell.row} = {value[:60]}")
        
        return references_found
        
    except Exception as e:
        print(f"ERROR: {e}")
        return []

def main():
    """Scan all Excel files for CAD references"""
    
    print("\n" + "="*70)
    print("EXCEL FILE REFERENCE SCANNER")
    print("Finding CAD file references that may need updating...")
    print("="*70)
    
    # Scan all tools
    tools = {
        'Hudson Certified': 'templates/hudson_certified',
        'Header Section Tool': 'templates/header_section_tool',
        'XCH Structure Tool': 'templates/xch_structure_tool',
        'Z Structure Tool': 'templates/z_structure_tool'
    }
    
    all_references = {}
    
    for tool_name, tool_path in tools.items():
        print(f"\n\n{'#'*70}")
        print(f"# {tool_name}")
        print(f"{'#'*70}")
        
        tool_path_obj = Path(tool_path)
        if not tool_path_obj.exists():
            print(f"  Path not found: {tool_path}")
            continue
        
        # Find all Excel files
        excel_files = list(tool_path_obj.rglob('*.xlsx')) + list(tool_path_obj.rglob('*.xls'))
        
        print(f"\nFound {len(excel_files)} Excel files")
        
        tool_references = []
        for excel_file in excel_files:
            refs = find_cad_references(excel_file)
            if refs:
                tool_references.extend(refs)
        
        all_references[tool_name] = {
            'excel_files_scanned': len(excel_files),
            'references_found': len(tool_references),
            'references': tool_references
        }
    
    # Print summary
    print(f"\n\n{'='*70}")
    print("SUMMARY")
    print(f"{'='*70}\n")
    
    total_refs = 0
    for tool_name, data in all_references.items():
        refs_count = data['references_found']
        total_refs += refs_count
        print(f"{tool_name}:")
        print(f"  Excel files: {data['excel_files_scanned']}")
        print(f"  References found: {refs_count}")
        
        if refs_count > 0:
            # Count by type
            types = {}
            for ref in data['references']:
                ref_type = ref.get('type', 'unknown')
                types[ref_type] = types.get(ref_type, 0) + 1
            
            for ref_type, count in types.items():
                print(f"    {ref_type}: {count}")
        print()
    
    print(f"\nTOTAL REFERENCES FOUND: {total_refs}")
    
    if total_refs > 0:
        print(f"\n{'='*70}")
        print("RECOMMENDATION:")
        print(f"{'='*70}")
        print("\nExcel files contain references to CAD files.")
        print("These may need updating to match renamed files:")
        print("  - JOBNO-* files now have HUD_ prefix")
        print("  - Z_* files now have ZST_ prefix")
        print("\nNext step: Create update script to fix references")
    else:
        print("\nNo direct file references found in Excel files.")
        print("Design tables may use embedded links or relative paths.")

if __name__ == "__main__":
    main()

