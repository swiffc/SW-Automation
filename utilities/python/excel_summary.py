"""
Excel Summary Utility
======================
A simple helper to read and summarize Excel mapping files used in SolidWorks automation.
This utility does NOT require SolidWorks and can be run standalone.

Usage:
    python excel_summary.py <path_to_excel_file>

Example:
    python excel_summary.py "../../templates/header_section_tool/Combined_/Drafting/Headers/000000_S01c-HCS.xlsx"
"""

import sys
import openpyxl
from pathlib import Path


def summarize_excel(file_path):
    """
    Read an Excel file and print a summary of sheets, dimensions, and key data.
    
    Args:
        file_path: Path to the Excel file to summarize
        
    Returns:
        dict: Summary information about the workbook
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    if file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        raise ValueError(f"File must be .xlsx or .xlsm format: {file_path}")
    
    print(f"\n{'='*60}")
    print(f"Excel Summary: {file_path.name}")
    print(f"{'='*60}\n")
    
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    summary = {
        'file': str(file_path),
        'sheets': []
    }
    
    print(f"Total sheets: {len(workbook.sheetnames)}\n")
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Count non-empty cells (sample first 100 rows for performance)
        non_empty_count = 0
        sample_rows = min(100, max_row)
        for row in sheet.iter_rows(min_row=1, max_row=sample_rows, max_col=max_col):
            non_empty_count += sum(1 for cell in row if cell.value is not None)
        
        sheet_info = {
            'name': sheet_name,
            'dimensions': f"{max_row} rows x {max_col} columns",
            'non_empty_cells': f"~{non_empty_count} (sampled)" if max_row > 100 else str(non_empty_count)
        }
        summary['sheets'].append(sheet_info)
        
        print(f"Sheet: '{sheet_name}'")
        print(f"  Dimensions: {max_row} rows x {max_col} columns")
        print(f"  Non-empty cells: {sheet_info['non_empty_cells']}")
        
        # Show first few non-empty cells as a sample
        print(f"  Sample data:")
        sample_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=min(10, max_row), max_col=min(5, max_col)):
            for cell in row:
                if cell.value is not None and sample_count < 5:
                    print(f"    {cell.coordinate}: {cell.value}")
                    sample_count += 1
            if sample_count >= 5:
                break
        print()
    
    workbook.close()
    
    print(f"{'='*60}\n")
    return summary


def main():
    """Main entry point for command-line usage."""
    if len(sys.argv) < 2:
        print("Usage: python excel_summary.py <path_to_excel_file>")
        print("\nExample:")
        print('  python excel_summary.py "../../templates/header_section_tool/Combined_/Drafting/Headers/000000_S01c-HCS.xlsx"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        summarize_excel(excel_file)
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
