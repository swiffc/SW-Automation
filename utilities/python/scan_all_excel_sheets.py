"""
Complete Excel Analysis - All Sheets, All Formulas
Scans S25140-Prego1.xlsm comprehensively
"""

import openpyxl
from openpyxl import load_workbook
import json
import sys

# Fix Unicode encoding for console output
sys.stdout.reconfigure(encoding='utf-8')

def scan_all_sheets(file_path):
    """Scan all sheets and extract formulas"""
    
    print(f"Loading Excel file: {file_path}")
    wb = load_workbook(file_path, data_only=False, keep_vba=True)
    
    all_sheets_data = {
        'filename': file_path.split('\\')[-1],
        'total_sheets': len(wb.sheetnames),
        'sheets': []
    }
    
    print(f"\nTotal Sheets Found: {len(wb.sheetnames)}")
    print("=" * 100)
    
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"\n[{idx}/{len(wb.sheetnames)}] Processing: {sheet_name}")
        
        try:
            ws = wb[sheet_name]
            
            sheet_info = {
                'index': idx,
                'name': sheet_name,
                'dimensions': str(ws.dimensions),
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'formulas': [],
                'important_cells': [],
                'cell_count': 0
            }
            
            # Count cells and formulas
            formula_count = 0
            data_count = 0
            
            for row in ws.iter_rows(max_row=min(ws.max_row, 200)):  # Limit to first 200 rows
                for cell in row:
                    if cell.value is not None:
                        data_count += 1
                        
                        # Check for formula
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_count += 1
                            
                            # Store formula info
                            formula_info = {
                                'cell': cell.coordinate,
                                'formula': str(cell.value)[:150],  # Limit length
                                'row': cell.row,
                                'col': cell.column
                            }
                            sheet_info['formulas'].append(formula_info)
                        
                        # Capture important cells (first few rows/columns)
                        elif cell.row <= 20 and cell.column <= 10:
                            try:
                                cell_value = str(cell.value)[:100]
                                if cell_value.strip():  # Only non-empty
                                    sheet_info['important_cells'].append({
                                        'cell': cell.coordinate,
                                        'value': cell_value
                                    })
                            except:
                                pass
            
            sheet_info['cell_count'] = data_count
            sheet_info['formula_count'] = formula_count
            
            print(f"  ✓ Cells: {data_count}, Formulas: {formula_count}")
            
            # Show first few formulas
            if formula_count > 0:
                print(f"  📐 Sample Formulas:")
                for formula in sheet_info['formulas'][:5]:
                    print(f"     {formula['cell']}: {formula['formula'][:80]}")
                if formula_count > 5:
                    print(f"     ... and {formula_count - 5} more formulas")
            
            all_sheets_data['sheets'].append(sheet_info)
            
        except Exception as e:
            print(f"  ❌ Error processing {sheet_name}: {str(e)[:100]}")
            continue
    
    # Save to JSON
    output_file = 'complete_excel_analysis.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_sheets_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'=' * 100}")
    print(f"✅ Complete analysis saved to: {output_file}")
    
    # Create summary report
    create_summary(all_sheets_data)
    
    return all_sheets_data


def create_summary(data):
    """Create readable summary report"""
    
    report_file = 'COMPLETE_EXCEL_SUMMARY.md'
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(f"# Complete Excel Analysis: {data['filename']}\n\n")
        f.write(f"**Total Sheets:** {data['total_sheets']}\n\n")
        f.write("---\n\n")
        
        # Summary table
        f.write("## Sheet Summary\n\n")
        f.write("| # | Sheet Name | Cells | Formulas | Dimensions |\n")
        f.write("|---|------------|-------|----------|------------|\n")
        
        total_formulas = 0
        for sheet in data['sheets']:
            f.write(f"| {sheet['index']} | {sheet['name']} | {sheet['cell_count']} | ")
            f.write(f"{sheet.get('formula_count', 0)} | {sheet['dimensions']} |\n")
            total_formulas += sheet.get('formula_count', 0)
        
        f.write(f"\n**Total Formulas Across All Sheets:** {total_formulas}\n\n")
        f.write("---\n\n")
        
        # Detail each sheet
        for sheet in data['sheets']:
            f.write(f"\n## {sheet['index']}. {sheet['name']}\n\n")
            f.write(f"- **Dimensions:** {sheet['dimensions']}\n")
            f.write(f"- **Data Cells:** {sheet['cell_count']}\n")
            f.write(f"- **Formulas:** {sheet.get('formula_count', 0)}\n\n")
            
            # Show formulas
            if sheet.get('formulas'):
                f.write(f"### Key Formulas:\n\n")
                for formula in sheet['formulas'][:20]:
                    f.write(f"- **{formula['cell']}**: `{formula['formula']}`\n")
                
                if len(sheet['formulas']) > 20:
                    f.write(f"\n*... and {len(sheet['formulas']) - 20} more formulas*\n")
            
            # Show important cells
            if sheet.get('important_cells') and len(sheet['important_cells']) > 0:
                f.write(f"\n### Sample Data:\n\n")
                for cell_info in sheet['important_cells'][:10]:
                    f.write(f"- **{cell_info['cell']}**: {cell_info['value']}\n")
            
            f.write("\n---\n")
    
    print(f"✅ Summary report saved to: {report_file}")


if __name__ == '__main__':
    file_path = r'c:\Users\DCornealius\CascadeProjects\Solidworks_Automation\S25140-Prego1.xlsm'
    
    try:
        print("Starting comprehensive Excel analysis...")
        print("This may take a few minutes...")
        print()
        
        data = scan_all_sheets(file_path)
        
        print("\n" + "=" * 100)
        print("✅ ANALYSIS COMPLETE!")
        print(f"   Total Sheets Scanned: {data['total_sheets']}")
        print(f"   Output Files:")
        print(f"     - complete_excel_analysis.json")
        print(f"     - COMPLETE_EXCEL_SUMMARY.md")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
