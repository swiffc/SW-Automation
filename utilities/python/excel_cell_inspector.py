"""
Excel Cell Inspector - Find Critical Engineering Parameters
Interactive tool to help map Excel cells to engineering parameters
"""

import openpyxl
from openpyxl import load_workbook
import json

def inspect_cells_interactive(file_path):
    """Interactive cell inspector"""
    
    print("=" * 100)
    print("EXCEL CELL INSPECTOR - Engineering Parameter Mapping Tool")
    print("=" * 100)
    print()
    print("Loading Excel file...")
    
    wb = load_workbook(file_path, data_only=False)
    
    # Focus on key sheets
    key_sheets = ['Inputs_Calcs', 'RAGU', 'Prego_to_Sw', 'Input']
    
    cell_mapping = {
        'bundle': {},
        'header': {},
        'common': {}
    }
    
    print("\nKey Sheets Available:")
    for idx, sheet in enumerate(key_sheets, 1):
        if sheet in wb.sheetnames:
            print(f"  {idx}. {sheet}")
    
    print("\n" + "=" * 100)
    print("Searching for Common Engineering Parameters...")
    print("=" * 100)
    
    # Search patterns for common parameters
    search_terms = {
        'Job Number': ['job', 'project', 'job number', 'job no'],
        'Bundle Width': ['bundle width', 'width', 'bundle w'],
        'Bundle Depth': ['bundle depth', 'depth', 'bundle d'],
        'Tube OD': ['tube od', 'tube diameter', 'od'],
        'Tube Length': ['tube length', 'length'],
        'Tube Wall': ['tube wall', 'wall thickness', 'thk'],
        'Tube Count': ['tube count', 'number of tubes', 'qty'],
        'Row Count': ['row', 'rows', 'tube row'],
        'Horizontal Pitch': ['horiz pitch', 'horizontal', 'pitch'],
        'Vertical Pitch': ['vert pitch', 'vertical'],
        'Header Width': ['header width', 'box width'],
        'Header Height': ['header height', 'box height'],
        'Header Length': ['header length', 'box length'],
        'Tubesheet Thickness': ['tubesheet', 'ts thick', 'tube sheet'],
        'Design Pressure': ['design pressure', 'design press', 'dp'],
        'MAWP': ['mawp', 'max allow'],
        'Material': ['material', 'matl', 'spec'],
    }
    
    found_cells = {}
    
    for sheet_name in key_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        print(f"\nSearching sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Search first 100 rows, 50 columns
        for row in ws.iter_rows(max_row=100, max_col=50):
            for cell in row:
                if cell.value is None:
                    continue
                
                cell_text = str(cell.value).lower()
                
                # Check each search term
                for param_name, keywords in search_terms.items():
                    for keyword in keywords:
                        if keyword in cell_text and len(cell_text) < 50:
                            # Found potential match
                            # Look for value in adjacent cells
                            adjacent_cells = []
                            
                            # Check right (same row, next column)
                            if cell.column < ws.max_column:
                                right_cell = ws.cell(cell.row, cell.column + 1)
                                if right_cell.value is not None:
                                    adjacent_cells.append({
                                        'location': 'right',
                                        'cell': right_cell.coordinate,
                                        'value': str(right_cell.value)[:50],
                                        'is_formula': isinstance(right_cell.value, str) and right_cell.value.startswith('=')
                                    })
                            
                            # Check below (next row, same column)
                            if cell.row < ws.max_row:
                                below_cell = ws.cell(cell.row + 1, cell.column)
                                if below_cell.value is not None:
                                    adjacent_cells.append({
                                        'location': 'below',
                                        'cell': below_cell.coordinate,
                                        'value': str(below_cell.value)[:50],
                                        'is_formula': isinstance(below_cell.value, str) and below_cell.value.startswith('=')
                                    })
                            
                            if adjacent_cells:
                                key = f"{sheet_name}_{param_name}"
                                if key not in found_cells:
                                    found_cells[key] = {
                                        'parameter': param_name,
                                        'sheet': sheet_name,
                                        'label_cell': cell.coordinate,
                                        'label_text': str(cell.value)[:50],
                                        'adjacent': adjacent_cells
                                    }
                                    print(f"  Found: {param_name} at {cell.coordinate}")
    
    # Save findings
    output_file = 'CELL_MAPPING_FINDINGS.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(found_cells, f, indent=2)
    
    print(f"\n{'=' * 100}")
    print(f"Found {len(found_cells)} potential parameter locations")
    print(f"Saved to: {output_file}")
    
    # Create readable report
    create_mapping_report(found_cells)
    
    return found_cells


def create_mapping_report(found_cells):
    """Create readable mapping report"""
    
    report_file = 'CELL_MAPPING_REPORT.md'
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("# Excel Cell Mapping Report\n\n")
        f.write("## Found Engineering Parameters\n\n")
        
        # Group by parameter
        by_param = {}
        for key, data in found_cells.items():
            param = data['parameter']
            if param not in by_param:
                by_param[param] = []
            by_param[param].append(data)
        
        for param in sorted(by_param.keys()):
            f.write(f"\n### {param}\n\n")
            
            for data in by_param[param]:
                f.write(f"**Sheet:** {data['sheet']}\n\n")
                f.write(f"- **Label Cell:** {data['label_cell']}\n")
                f.write(f"- **Label Text:** {data['label_text']}\n")
                
                if data['adjacent']:
                    f.write(f"- **Value Cells:**\n")
                    for adj in data['adjacent']:
                        formula_note = " (FORMULA)" if adj['is_formula'] else ""
                        f.write(f"  - {adj['location'].upper()}: {adj['cell']} = `{adj['value']}`{formula_note}\n")
                
                f.write("\n")
        
        f.write("\n---\n\n")
        f.write("## Recommended Mapping\n\n")
        f.write("```csharp\n")
        f.write("// Bundle Parameters\n")
        for param in sorted(by_param.keys()):
            if 'bundle' in param.lower() or 'tube' in param.lower():
                if by_param[param]:
                    data = by_param[param][0]
                    if data['adjacent']:
                        cell_ref = data['adjacent'][0]['cell']
                        sheet = data['sheet']
                        param_safe = param.replace(' ', '')
                        f.write(f"config.{param_safe} = ReadCell(\"{sheet}\", \"{cell_ref}\");\n")
        
        f.write("\n// Header Parameters\n")
        for param in sorted(by_param.keys()):
            if 'header' in param.lower() or 'tubesheet' in param.lower():
                if by_param[param]:
                    data = by_param[param][0]
                    if data['adjacent']:
                        cell_ref = data['adjacent'][0]['cell']
                        sheet = data['sheet']
                        param_safe = param.replace(' ', '')
                        f.write(f"config.{param_safe} = ReadCell(\"{sheet}\", \"{cell_ref}\");\n")
        
        f.write("```\n")
    
    print(f"Mapping report saved to: {report_file}")


if __name__ == '__main__':
    file_path = r'c:\Users\DCornealius\CascadeProjects\Solidworks_Automation\S25140-Prego1.xlsm'
    
    try:
        found = inspect_cells_interactive(file_path)
        
        print("\n" + "=" * 100)
        print("SUCCESS: Cell inspection complete!")
        print("=" * 100)
        print("\nNext steps:")
        print("1. Review CELL_MAPPING_REPORT.md")
        print("2. Verify the cell addresses are correct")
        print("3. Update ExcelTemplateImporter with exact cell mappings")
        print("=" * 100)
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
