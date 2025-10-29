"""
AI-Powered Excel Validator
===========================
Validates changes to Excel configuration files (HCS, SCS, etc.) and checks for:
- Breaking changes to expected column schemas
- Missing updates to C# code that references Excel columns
- Missing documentation updates

Setup:
1. Copy .env.example to .env
2. Add your OpenAI API key to .env
3. Run: pip install -r requirements.txt

Usage:
    # Validate a modified Excel file
    python ai_excel_validator.py "path/to/modified_HCS.xlsx"
    
    # Compare old vs new version
    python ai_excel_validator.py "path/to/old.xlsx" "path/to/new.xlsx"
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple
import json

try:
    from openai import OpenAI
    from dotenv import load_dotenv
    import openpyxl
except ImportError:
    print("Error: Required packages not installed.")
    print("Run: pip install -r requirements.txt")
    sys.exit(1)


class ExcelValidator:
    """AI-powered validator for Excel configuration files."""
    
    def __init__(self):
        """Initialize the validator."""
        load_dotenv()
        
        self.api_key = os.getenv('OPENAI_API_KEY')
        if not self.api_key or self.api_key == 'your-openai-api-key-here':
            raise ValueError(
                "OpenAI API key not found!\n"
                "1. Copy .env.example to .env\n"
                "2. Add your OpenAI API key to .env"
            )
        
        self.client = OpenAI(api_key=self.api_key)
        self.model = os.getenv('OPENAI_MODEL', 'gpt-4-turbo-preview')
        
        # Get repo root
        self.repo_root = Path(__file__).parent.parent.parent
    
    def extract_schema(self, excel_path: Path) -> Dict:
        """Extract schema information from an Excel file."""
        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        
        schema = {
            'file': str(excel_path.name),
            'sheets': {}
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get headers (first row)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            # Get a few sample rows
            sample_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True)):
                if any(cell is not None for cell in row):
                    sample_rows.append([str(cell) if cell is not None else '' for cell in row[:len(headers)]])
            
            schema['sheets'][sheet_name] = {
                'columns': headers,
                'row_count': sheet.max_row,
                'sample_data': sample_rows
            }
        
        workbook.close()
        return schema
    
    def validate_changes(self, file_path: Path, old_schema: Dict = None) -> Dict:
        """
        Validate Excel file changes using AI.
        
        Args:
            file_path: Path to Excel file to validate
            old_schema: Optional previous schema for comparison
            
        Returns:
            Validation results with warnings and recommendations
        """
        print(f"\n{'='*60}")
        print(f"Validating: {file_path.name}")
        print(f"{'='*60}\n")
        
        # Extract current schema
        current_schema = self.extract_schema(file_path)
        
        # Load integration documentation
        integration_docs = self._load_integration_docs(file_path)
        
        # Build validation prompt
        prompt = self._build_validation_prompt(
            current_schema, 
            old_schema, 
            integration_docs,
            file_path
        )
        
        # Call OpenAI API
        print("?? Analyzing with AI...\n")
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are an expert at validating Excel configuration files for SolidWorks automation. Analyze changes for safety and completeness."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.3  # Lower temperature for more consistent validation
            )
            
            analysis = response.choices[0].message.content
            
            print(analysis)
            print(f"\n{'='*60}\n")
            
            return {
                'file': str(file_path),
                'schema': current_schema,
                'analysis': analysis
            }
            
        except Exception as e:
            print(f"? Error calling OpenAI API: {e}")
            return {'error': str(e)}
    
    def _load_integration_docs(self, file_path: Path) -> str:
        """Load relevant integration documentation."""
        docs = []
        
        # Determine which integration guide to load based on file name
        if 'HCS' in file_path.name.upper():
            doc_path = self.repo_root / 'docs' / 'HEADER_SECTION_TOOL_INTEGRATION.md'
            if doc_path.exists():
                try:
                    docs.append(doc_path.read_text(encoding='utf-8'))
                except UnicodeDecodeError:
                    docs.append(doc_path.read_text(encoding='utf-8', errors='ignore'))
        
        if 'XCH' in file_path.name.upper() or 'SCS' in file_path.name.upper():
            doc_path = self.repo_root / 'docs' / 'XCH_STRUCTURE_TOOL_INTEGRATION.md'
            if doc_path.exists():
                try:
                    docs.append(doc_path.read_text(encoding='utf-8'))
                except UnicodeDecodeError:
                    docs.append(doc_path.read_text(encoding='utf-8', errors='ignore'))
        
        # Always load AGENTS.md for safety rules
        agents_path = self.repo_root / 'AGENTS.md'
        if agents_path.exists():
            try:
                docs.append(agents_path.read_text(encoding='utf-8')[:2000])
            except UnicodeDecodeError:
                docs.append(agents_path.read_text(encoding='utf-8', errors='ignore')[:2000])
        
        return '\n\n'.join(docs) if docs else "No integration docs found."
    
    def _build_validation_prompt(self, current_schema: Dict, old_schema: Dict, 
                                  integration_docs: str, file_path: Path) -> str:
        """Build the validation prompt for AI analysis."""
        prompt = f"""Validate this Excel configuration file for a SolidWorks automation tool.

FILE: {current_schema['file']}
TYPE: {"Header Section Config (HCS)" if "HCS" in file_path.name.upper() else "Structure Config"}

CURRENT SCHEMA:
{json.dumps(current_schema, indent=2)}

"""
        
        if old_schema:
            prompt += f"""
PREVIOUS SCHEMA (for comparison):
{json.dumps(old_schema, indent=2)}

"""
        
        prompt += f"""
INTEGRATION DOCUMENTATION (for reference):
{integration_docs[:3000]}  # First 3000 chars

VALIDATION TASKS:
1. **Schema Analysis**: List all column names and their apparent purposes
2. **Breaking Changes**: Identify any columns removed or renamed (if comparing to old schema)
3. **C# Code Impact**: Which C# files likely reference these Excel columns?
   - Check: macros/csharp/Solidworks-Automation/**/
   - Search for: column names as string literals or property names
4. **Documentation Updates**: Does docs/ need updates for schema changes?
5. **Safety Checklist**: What manual verification steps are needed?

OUTPUT FORMAT:
## ? Schema Summary
[List columns and purposes]

## ?? Breaking Changes
[List any breaking changes, or "None detected"]

## ?? Required Code Updates
[List C# files that likely need updates]

## ?? Documentation Updates
[List docs that need updates]

## ? Safety Checklist
[Manual verification steps before merge]
"""
        
        return prompt


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python ai_excel_validator.py <excel_file>")
        print("  python ai_excel_validator.py <old_file> <new_file>")
        print("\nExample:")
        print('  python ai_excel_validator.py "../../templates/header_section_tool/Combined_/Drafting/Headers/000000_S01c-HCS.xlsx"')
        sys.exit(1)
    
    try:
        validator = ExcelValidator()
        
        file_path = Path(sys.argv[1])
        if not file_path.exists():
            print(f"? Error: File not found: {file_path}")
            sys.exit(1)
        
        # If two files provided, compare them
        old_schema = None
        if len(sys.argv) > 2:
            old_path = Path(sys.argv[2])
            if old_path.exists():
                print(f"?? Extracting old schema from: {old_path.name}")
                old_schema = validator.extract_schema(old_path)
        
        # Validate
        result = validator.validate_changes(file_path, old_schema)
        
        # Save results to JSON
        output_file = Path('excel_validation_results.json')
        with open(output_file, 'w') as f:
            json.dump(result, f, indent=2)
        print(f"?? Results saved to: {output_file}")
        
    except ValueError as e:
        print(f"\n? Configuration Error:\n{e}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n? Unexpected Error: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
