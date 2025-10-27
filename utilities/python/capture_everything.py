"""
COMPLETE Excel Capture - Everything from S25140-Prego1.xlsm
All sheets, all cells, all formulas, all data
"""

import openpyxl
from openpyxl import load_workbook
import json
import sys

def capture_everything(file_path):
    """Capture EVERYTHING from Excel file"""
    
    print(f"Loading: {file_path}")
    wb = load_workbook(file_path, data_only=False, keep_vba=True)
    
    complete_data = {
        'filename': 'S25140-Prego1.xlsm',
        'total_sheets': len(wb.sheetnames),
        'sheet_names': wb.sheetnames,
        'sheets': {}
    }
    
    print(f"\nCap Capturing {len(wb.sheetnames)} sheets...")
    print("=" * 100)
    
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"\n[{idx}/{len(wb.sheetnames)}] {sheet_name}")
        
        try:
            ws = wb[sheet_name]
            
            sheet_data = {
                'index': idx,
                'name': sheet_name,
                'dimensions': str(ws.dimensions),
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'cells': {},
                'formulas': {},
                'merged_cells': [str(r) for r in ws.merged_cells.ranges],
                'statistics': {
                    'total_cells': 0,
                    'formula_cells': 0,
                    'value_cells': 0,
                    'empty_cells': 0
                }
            }
            
            # Capture ALL cells
            for row in ws.iter_rows():
                for cell in row:
                    coord = cell.coordinate
                    
                    if cell.value is not None:
                        cell_info = {
                            'value': str(cell.value),
                            'data_type': cell.data_type,
                            'row': cell.row,
                            'column': cell.column,
                            'number_format': cell.number_format
                        }
                        
                        # Check if it's a formula
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            sheet_data['formulas'][coord] = cell.value
                            cell_info['is_formula'] = True
                            sheet_data['statistics']['formula_cells'] += 1
                        else:
                            sheet_data['statistics']['value_cells'] += 1
                        
                        sheet_data['cells'][coord] = cell_info
                        sheet_data['statistics']['total_cells'] += 1
            
            complete_data['sheets'][sheet_name] = sheet_data
            
            print(f"  OK Captured {sheet_data['statistics']['total_cells']} cells")
            print(f"    - Formulas: {sheet_data['statistics']['formula_cells']}")
            print(f"    - Values: {sheet_data['statistics']['value_cells']}")
            
        except Exception as e:
            print(f"  ERROR: {str(e)[:100]}")
            complete_data['sheets'][sheet_name] = {'error': str(e)}
    
    # Save complete data
    print(f"\n{'=' * 100}")
    print("Saving complete capture...")
    
    output_file = 'COMPLETE_EXCEL_CAPTURE.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(complete_data, f, indent=2, ensure_ascii=False)
    
    print(f"SUCCESS: Complete capture saved to: {output_file}")
    
    # Create readable summary
    create_detailed_report(complete_data)
    
    return complete_data


def create_detailed_report(data):
    """Create detailed markdown report"""
    
    report_file = 'COMPLETE_EXCEL_REPORT.md'
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(f"# Complete Excel Capture: {data['filename']}\n\n")
        f.write(f"**Generated:** {data.get('timestamp', 'N/A')}\n\n")
        f.write(f"**Total Sheets:** {data['total_sheets']}\n\n")
        
        # Quick stats
        total_formulas = sum(
            sheet.get('statistics', {}).get('formula_cells', 0) 
            for sheet in data['sheets'].values() 
            if 'statistics' in sheet
        )
        total_cells = sum(
            sheet.get('statistics', {}).get('total_cells', 0)
            for sheet in data['sheets'].values()
            if 'statistics' in sheet
        )
        
        f.write(f"**Total Cells:** {total_cells:,}\n")
        f.write(f"**Total Formulas:** {total_formulas:,}\n\n")
        f.write("---\n\n")
        
        # Sheet index
        f.write("## Sheet Index\n\n")
        for idx, name in enumerate(data['sheet_names'], 1):
            sheet = data['sheets'].get(name, {})
            stats = sheet.get('statistics', {})
            f.write(f"{idx}. **{name}** - {stats.get('total_cells', 0)} cells, ")
            f.write(f"{stats.get('formula_cells', 0)} formulas\n")
        
        f.write("\n---\n\n")
        
        # Detailed sheet info
        f.write("## Detailed Sheet Information\n\n")
        
        for sheet_name in data['sheet_names']:
            sheet = data['sheets'].get(sheet_name, {})
            
            if 'error' in sheet:
                f.write(f"### {sheet_name}\n\n")
                f.write(f"❌ Error: {sheet['error']}\n\n")
                continue
            
            f.write(f"### {sheet.get('index', '?')}. {sheet_name}\n\n")
            
            stats = sheet.get('statistics', {})
            f.write(f"- **Dimensions:** {sheet.get('dimensions', 'N/A')}\n")
            f.write(f"- **Max Row:** {sheet.get('max_row', 0):,}\n")
            f.write(f"- **Max Column:** {sheet.get('max_column', 0)}\n")
            f.write(f"- **Total Cells:** {stats.get('total_cells', 0):,}\n")
            f.write(f"- **Formula Cells:** {stats.get('formula_cells', 0):,}\n")
            f.write(f"- **Value Cells:** {stats.get('value_cells', 0):,}\n\n")
            
            # Show formulas if present
            formulas = sheet.get('formulas', {})
            if formulas:
                f.write(f"#### Formulas ({len(formulas)})\n\n")
                
                # Show first 30 formulas
                for idx, (cell, formula) in enumerate(list(formulas.items())[:30]):
                    f.write(f"- **{cell}**: `{formula[:150]}`\n")
                
                if len(formulas) > 30:
                    f.write(f"\n*... and {len(formulas) - 30} more formulas*\n")
            
            # Show sample cells
            cells = sheet.get('cells', {})
            if cells:
                non_formula_cells = {k: v for k, v in list(cells.items())[:20] 
                                    if not v.get('is_formula', False)}
                
                if non_formula_cells:
                    f.write(f"\n#### Sample Data (first 20 values)\n\n")
                    for cell, info in non_formula_cells.items():
                        value = info['value'][:100]
                        f.write(f"- **{cell}**: {value}\n")
            
            # Merged cells
            merged = sheet.get('merged_cells', [])
            if merged:
                f.write(f"\n#### Merged Cells ({len(merged)})\n\n")
                for merge in merged[:10]:
                    f.write(f"- {merge}\n")
                if len(merged) > 10:
                    f.write(f"*... and {len(merged) - 10} more*\n")
            
            f.write("\n---\n\n")
    
    print(f"SUCCESS: Detailed report saved to: {report_file}")


if __name__ == '__main__':
    import datetime
    
    file_path = r'c:\Users\DCornealius\CascadeProjects\Solidworks_Automation\S25140-Prego1.xlsm'
    
    try:
        print("=" * 100)
        print("COMPLETE EXCEL CAPTURE - EVERYTHING")
        print("=" * 100)
        print("\nThis will capture:")
        print("  - All 73 sheets")
        print("  - Every single cell")
        print("  - All formulas")
        print("  - All values")
        print("  - Merged cells")
        print("  - Formatting info")
        print("\nThis may take 2-3 minutes...")
        print()
        
        start_time = datetime.datetime.now()
        
        data = capture_everything(file_path)
        data['timestamp'] = start_time.isoformat()
        
        end_time = datetime.datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        print("\n" + "=" * 100)
        print("SUCCESS: CAPTURE COMPLETE!")
        print(f"   Duration: {duration:.1f} seconds")
        print(f"   Sheets: {data['total_sheets']}")
        print(f"\n   Output Files:")
        print(f"     1. COMPLETE_EXCEL_CAPTURE.json  (Full data)")
        print(f"     2. COMPLETE_EXCEL_REPORT.md     (Readable report)")
        print("=" * 100)
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
