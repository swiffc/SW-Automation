# -*- coding: utf-8 -*-
"""Quick check of master config files only"""

import openpyxl
from pathlib import Path

# Just check the 2 most important files
files = [
    "templates/header_section_tool/Combined_/Drafting/Headers/000000_S01c-HCS.xlsx",
    "templates/xch_structure_tool/XCH Cooler/XCH_SCS.xlsx",
]

print("\nQuick Excel Reference Check")
print("="*60)

for filepath in files:
    path = Path(filepath)
    if not path.exists():
        print(f"\nNot found: {filepath}")
        continue
    
    print(f"\n{path.name}")
    print("-"*60)
    
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        
        refs_found = 0
        for sheet_name in wb.sheetnames[:3]:  # Only first 3 sheets
            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name} ({ws.max_row} rows)")
            
            # Quick scan of first 20 rows only
            for row in range(1, min(21, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    value = str(cell.value) if cell.value else ""
                    
                    if any(ext in value.upper() for ext in ['.SLDASM', '.SLDPRT', '.SLDDRW']):
                        print(f"    Row {row}, Col {col}: {value[:50]}")
                        refs_found += 1
        
        print(f"  References found: {refs_found}")
        wb.close()
        
    except Exception as e:
        print(f"  Error: {e}")

print("\n" + "="*60)
print("? Quick scan complete!")
print("\nNOTE: Excel design tables typically use embedded links")
print("that SolidWorks manages automatically, not direct file paths.")

