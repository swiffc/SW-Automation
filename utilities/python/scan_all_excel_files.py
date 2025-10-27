"""
Complete Excel File Scanner for SolidWorks Automation Project
Scans entire project and categorizes all Excel files
"""

import os
from pathlib import Path
import openpyxl
from collections import defaultdict

def get_file_size_mb(filepath):
    """Get file size in MB"""
    try:
        size_bytes = os.path.getsize(filepath)
        return round(size_bytes / (1024 * 1024), 2)
    except:
        return 0

def scan_excel_files(root_path):
    """Scan all Excel files in project"""
    
    excel_extensions = ('.xlsx', '.xls', '.xlsm')
    all_files = []
    
    print("\n" + "="*80)
    print("COMPLETE EXCEL FILE SCAN - SolidWorks Automation Project")
    print("="*80)
    
    # Walk through all directories
    for root, dirs, files in os.walk(root_path):
        for file in files:
            if file.endswith(excel_extensions):
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, root_path)
                
                file_info = {
                    'name': file,
                    'path': rel_path,
                    'full_path': full_path,
                    'ext': os.path.splitext(file)[1],
                    'size_mb': get_file_size_mb(full_path),
                    'category': categorize_file(rel_path)
                }
                all_files.append(file_info)
    
    return all_files

def categorize_file(rel_path):
    """Categorize Excel file based on path"""
    path_lower = rel_path.lower()
    
    if 'header_section_tool' in path_lower:
        if 'combined' in path_lower:
            return 'Header Section Tool - Combined'
        elif 'single' in path_lower:
            return 'Header Section Tool - Single'
        elif 'hailguard' in path_lower:
            return 'Header Section Tool - Hailguard'
        elif 'steam' in path_lower:
            return 'Header Section Tool - Steam Coil'
        else:
            return 'Header Section Tool'
    elif 'xch_structure' in path_lower or 'xch cooler' in path_lower:
        return 'XCH Structure Tool'
    elif 'z_structure' in path_lower or 'z cooler' in path_lower:
        return 'Z Structure Tool'
    elif 'standards' in path_lower:
        return 'Standards & Reference'
    elif 'hudson' in path_lower:
        return 'Hudson Certified'
    elif 'bundle' in path_lower:
        return 'Bundle'
    elif 'header' in path_lower and 'section' not in path_lower:
        return 'Header (Simple)'
    elif 'hood' in path_lower:
        return 'Hood'
    elif 'machinery' in path_lower:
        return 'Machinery Mount'
    elif 'plenum' in path_lower:
        return 'Plenum'
    elif 'structure' in path_lower and 'xch' not in path_lower and 'z_' not in path_lower:
        return 'Structure'
    elif 'walkway' in path_lower:
        return 'Walkway'
    else:
        return 'Other'

def analyze_excel_content(filepath, filename):
    """Analyze Excel file content"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        sheet_count = len(wb.sheetnames)
        
        # Get first sheet info
        first_sheet = wb[wb.sheetnames[0]]
        rows = first_sheet.max_row
        cols = first_sheet.max_column
        
        # Count formulas in first sheet
        formula_count = 0
        for row in first_sheet.iter_rows(max_row=min(100, rows)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        return {
            'sheets': sheet_count,
            'sheet_names': wb.sheetnames,
            'first_sheet_rows': rows,
            'first_sheet_cols': cols,
            'formulas_found': formula_count,
            'readable': True
        }
    except Exception as e:
        return {
            'sheets': 0,
            'sheet_names': [],
            'first_sheet_rows': 0,
            'first_sheet_cols': 0,
            'formulas_found': 0,
            'readable': False,
            'error': str(e)
        }

def main():
    root_path = Path('.')
    
    # Scan all files
    all_files = scan_excel_files(root_path)
    
    # Group by category
    by_category = defaultdict(list)
    for file_info in all_files:
        by_category[file_info['category']].append(file_info)
    
    # Print summary
    print(f"\n{'='*80}")
    print(f"SCAN COMPLETE")
    print(f"{'='*80}\n")
    print(f"Total Excel Files Found: {len(all_files)}")
    print(f"Total Size: {sum(f['size_mb'] for f in all_files):.2f} MB")
    print(f"Categories: {len(by_category)}")
    
    # Print by category
    print(f"\n{'='*80}")
    print("FILES BY CATEGORY")
    print(f"{'='*80}\n")
    
    for category in sorted(by_category.keys()):
        files = by_category[category]
        total_size = sum(f['size_mb'] for f in files)
        
        print(f"\n📁 {category} ({len(files)} files, {total_size:.2f} MB)")
        print("-" * 80)
        
        for idx, file_info in enumerate(sorted(files, key=lambda x: x['name']), 1):
            print(f"  {idx}. {file_info['name']}")
            print(f"     Path: {file_info['path']}")
            print(f"     Size: {file_info['size_mb']} MB")
            print()
    
    # Detailed analysis for key files
    print(f"\n{'='*80}")
    print("DETAILED ANALYSIS - KEY CONFIGURATION FILES")
    print(f"{'='*80}\n")
    
    key_files = [
        'templates/header_section_tool/Combined_/Drafting/Headers/000000_S01c-HCS.xlsx',
        'templates/header_section_tool/Single_/Drafting/Headers/000000_S03-HCS.xlsx',
        'templates/xch_structure_tool/XCH Cooler/XCH_SCS.xlsx',
    ]
    
    for rel_path in key_files:
        full_path = Path(rel_path)
        if full_path.exists():
            print(f"\n📊 {full_path.name}")
            print("-" * 80)
            analysis = analyze_excel_content(full_path, full_path.name)
            
            if analysis['readable']:
                print(f"  Sheets: {analysis['sheets']}")
                print(f"  Sheet Names: {', '.join(analysis['sheet_names'])}")
                print(f"  First Sheet: {analysis['first_sheet_rows']} rows × {analysis['first_sheet_cols']} cols")
                print(f"  Formulas Found: {analysis['formulas_found']}")
            else:
                print(f"  ⚠️ Could not read file: {analysis.get('error', 'Unknown error')}")
    
    # Save detailed report
    output_file = "docs/COMPLETE_EXCEL_SCAN_REPORT.txt"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("COMPLETE EXCEL FILE SCAN REPORT\n")
        f.write("SolidWorks Automation Project\n")
        f.write("="*80 + "\n\n")
        
        f.write(f"Total Files: {len(all_files)}\n")
        f.write(f"Total Size: {sum(file_info['size_mb'] for file_info in all_files):.2f} MB\n")
        f.write(f"Categories: {len(by_category)}\n\n")
        
        f.write("="*80 + "\n")
        f.write("COMPLETE FILE LIST\n")
        f.write("="*80 + "\n\n")
        
        for idx, file_info in enumerate(sorted(all_files, key=lambda x: x['path']), 1):
            f.write(f"{idx}. {file_info['name']}\n")
            f.write(f"   Category: {file_info['category']}\n")
            f.write(f"   Path: {file_info['path']}\n")
            f.write(f"   Extension: {file_info['ext']}\n")
            f.write(f"   Size: {file_info['size_mb']} MB\n")
            f.write("\n")
    
    print(f"\n{'='*80}")
    print(f"📄 Detailed report saved to: {output_file}")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    main()
